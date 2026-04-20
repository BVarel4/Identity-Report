"""Reporting layer for JSON, CSV and Excel outputs produced by the pipeline."""

from datetime import datetime
from pathlib import Path
import re

from actionability import (
    actionability_sort_rank,
    enrich_parsed_risks_with_actionability,
    priority_sort_rank,
)
from analytics import (
    build_entity_correlation_rows,
    build_executive_analytics,
    enrich_parsed_risks_with_entity_correlation,
)
from audit import AuditCollector
from config import PROJECT_SIGNATURE, Settings
from utils import (
    ensure_output_dir,
    get_timestamp,
    print_section,
    print_status,
    remove_dir_if_exists,
    remove_file_if_exists,
    save_csv,
    save_json,
)

RISK_INVENTORY_FIELDS = ["risk_type", "count", "percentage"]
PARSER_INVENTORY_FIELDS = [
    "risk_type",
    "risk_title",
    "risk_family",
    "count",
    "percentage",
    "selected_parser",
]
PARSED_RISK_FIELDS = [
    "entity",
    "secondary_identifier",
    "entity_type",
    "severity",
    "review_priority",
    "actionability_level",
    "actionability_note",
    "entity_risk_type_count",
    "entity_total_findings",
    "entity_risk_overview",
    "entity_family_overview",
    "risk_type",
    "risk_title",
    "risk_family",
    "risk_typename",
    "structure_profile",
    "context_summary",
    "technical_observation",
    "likely_impact",
    "evidence_available",
    "recommended_action",
    "detail",
    "related_entity",
    "related_falcon_link",
    "parser_used",
    "parse_status",
    "falcon_link",
]
PARSED_RISK_EXCEL_FIELDS = [
    "entity",
    "secondary_identifier",
    "entity_type",
    "severity",
    "review_priority",
    "actionability_level",
    "actionability_note",
    "entity_risk_type_count",
    "entity_total_findings",
    "entity_risk_overview",
    "risk_type",
    "risk_title",
    "risk_family",
    "structure_profile",
    "context_summary",
    "technical_observation",
    "likely_impact",
    "evidence_available",
    "recommended_action",
    "falcon_link",
]
ENTITY_CORRELATION_FIELDS = [
    "entity",
    "secondary_identifier",
    "entity_type",
    "highest_severity",
    "total_findings",
    "distinct_risk_types",
    "distinct_families",
    "risk_titles",
    "risk_families",
    "correlation_note",
    "falcon_link",
]
ATTACK_PATH_FIELDS = [
    "source_entity",
    "related_entity",
    "risk_title",
    "attack_summary",
    "attack_techniques",
    "attack_outcome",
    "attack_stage_count",
    "attack_stages",
    "attack_chain",
    "severity",
    "source_link",
    "related_link",
]
UNKNOWN_RISK_FIELDS = ["risk_type", "count", "action_required"]
ERROR_FIELDS = [
    "issue_type",
    "risk_type",
    "parser_name",
    "entity_name",
    "secondary_name",
    "entity_type",
    "severity",
    "message",
]
RAW_SAMPLE_FIELDS = [
    "risk_type",
    "sample_number",
    "sample_category",
    "parser_name",
    "entity_name",
    "secondary_name",
    "entity_type",
    "severity",
]

FIELD_LABELS = {
    "risk_type": "Tipo de riesgo",
    "risk_title": "Riesgo",
    "risk_family": "Familia",
    "count": "Cantidad",
    "percentage": "Porcentaje",
    "selected_parser": "Parser seleccionado",
    "entity": "Entidad",
    "secondary_identifier": "Identificador secundario",
    "entity_type": "Tipo de entidad",
    "severity": "Severidad de entidad",
    "review_priority": "Prioridad de revision",
    "actionability_level": "Nivel de accionabilidad",
    "actionability_note": "Siguiente paso sugerido",
    "highest_severity": "Severidad mas alta de entidad",
    "entity_risk_type_count": "Tipos de riesgo en la entidad",
    "entity_total_findings": "Hallazgos totales en la entidad",
    "total_findings": "Hallazgos totales",
    "entity_risk_overview": "Correlacion de entidad",
    "entity_family_overview": "Familias en la entidad",
    "distinct_risk_types": "Tipos de riesgo distintos",
    "distinct_families": "Familias distintas",
    "risk_titles": "Riesgos correlacionados",
    "risk_families": "Familias correlacionadas",
    "correlation_note": "Interpretacion de correlacion",
    "risk_typename": "Tipo GraphQL",
    "structure_profile": "Perfil estructural",
    "context_summary": "Contexto del hallazgo",
    "technical_observation": "Observacion tecnica",
    "likely_impact": "Impacto probable",
    "evidence_available": "Evidencia disponible",
    "detail": "Detalle tecnico",
    "recommended_action": "Accion recomendada",
    "related_entity": "Entidad relacionada",
    "related_falcon_link": "Enlace Falcon relacionado",
    "parser_used": "Parser usado",
    "parse_status": "Estado de parseo",
    "falcon_link": "Enlace Falcon",
    "source_entity": "Entidad origen",
    "attack_summary": "Resumen de la ruta",
    "attack_techniques": "Tecnicas observadas",
    "attack_outcome": "Resultado esperado",
    "attack_stage_count": "Cantidad de etapas",
    "attack_stages": "Ruta por etapas",
    "attack_chain": "Cadena compacta",
    "source_link": "Enlace origen",
    "related_link": "Enlace relacionado",
    "action_required": "Accion requerida",
    "issue_type": "Tipo de hallazgo",
    "parser_name": "Nombre del parser",
    "entity_name": "Nombre de entidad",
    "secondary_name": "Identificador secundario",
    "message": "Mensaje",
}

SHEET_TITLES = {
    "executive": "Resumen Ejecutivo",
    "risk_summary": "Resumen Riesgos",
    "parsed_risks": "Riesgos Parseados",
    "attack_paths": "Rutas de Ataque",
    "entity_correlation": "Correlacion Entidades",
    "lifecycle_correlation": "Correlacion Ciclo Vida",
    "unknown": "Auditoria - Sin Parser",
    "errors": "Auditoria - Errores",
}

RISK_GROUP_SHEETS = {
    "Credenciales Identidad": {
        "Password Hygiene",
        "Identity Hygiene",
        "Identity Correlation",
        "Credential Abuse",
        "Kerberos Exposure",
        "Certificate Exposure",
    },
    "Ciclo Vida Acceso": {
        "Account Lifecycle",
        "Access Change",
        "Behavioral Anomaly",
    },
    "Endpoint Directorio": {
        "Endpoint Hardening",
        "Endpoint Exposure",
        "Endpoint Posture",
        "Directory Hardening",
    },
    "Amenaza Privilegios": {
        "Threat Activity",
        "Privilege Exposure",
        "Lateral Movement",
    },
}

PERCENTAGE_FIELDS = {"percentage"}
COUNT_FIELDS = {
    "count",
    "attack_stage_count",
    "entity_risk_type_count",
    "entity_total_findings",
    "total_findings",
    "distinct_risk_types",
    "distinct_families",
}
URL_FIELDS = {"falcon_link", "related_falcon_link", "source_link", "related_link"}


def sort_parsed_risks(rows: list[dict]) -> list[dict]:
    severity_order = {
        "CRITICAL": 0,
        "HIGH": 1,
        "MEDIUM": 2,
        "LOW": 3,
        "INFO": 4,
        "": 5,
    }
    return sorted(
        rows,
        key=lambda row: (
            priority_sort_rank(str(row.get("review_priority", "") or "")),
            actionability_sort_rank(str(row.get("actionability_level", "") or "")),
            str(row.get("risk_family", "") or ""),
            str(row.get("risk_title", "") or ""),
            severity_order.get(str(row.get("severity", "") or "").upper(), 5),
            str(row.get("entity", "") or ""),
        ),
    )


def build_grouped_parsed_risk_sheets(parsed_risks: list[dict]) -> dict[str, list[dict]]:
    grouped = {title: [] for title in RISK_GROUP_SHEETS}
    others: list[dict] = []

    for row in parsed_risks:
        family = str(row.get("risk_family", "") or "")
        target_title = None
        for title, families in RISK_GROUP_SHEETS.items():
            if family in families:
                target_title = title
                break

        if target_title:
            grouped[target_title].append(row)
        else:
            others.append(row)

    result = {}
    for title, rows in grouped.items():
        if rows:
            result[title] = sort_parsed_risks(rows)

    if others:
        result["Otros Riesgos"] = sort_parsed_risks(others)

    return result


def sanitize_filename_component(value: str) -> str:
    normalized = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip())
    normalized = normalized.strip("._")
    return normalized or "unknown"


def apply_artifact_retention(
    artifact_mode: str,
    run_dir: Path,
    output_files: dict,
) -> tuple[dict, list[str]]:
    notes = []

    if artifact_mode == "debug":
        output_files["artifact_mode"] = "debug"
        output_files["run_artifacts_dir"] = str(run_dir)
        notes.append("Modo debug: se conservaron todos los artefactos de trabajo.")
        return output_files, notes

    if artifact_mode == "standard":
        sensitive_keys = [
            "discovery_raw_json",
            "detail_raw_json",
            "raw_samples_json",
        ]
        removed = 0
        for key in sensitive_keys:
            path = output_files.get(key)
            if path and remove_file_if_exists(path):
                removed += 1
        for key in sensitive_keys:
            output_files.pop(key, None)

        output_files["artifact_mode"] = "standard"
        output_files["run_artifacts_dir"] = str(run_dir)
        notes.append(
            "Modo standard: se eliminaron artefactos raw sensibles y se conservaron CSV tecnicos."
        )
        notes.append(f"Archivos sensibles purgados: {removed}")
        return output_files, notes

    removed_dir = remove_dir_if_exists(run_dir)
    retained_keys = {
        "technical_excel": output_files.get("technical_excel", ""),
        "artifact_mode": "final_only",
    }
    if removed_dir:
        notes.append(
            "Modo final_only: se eliminaron los artefactos intermedios y solo se conservo el reporte final."
        )
        notes.append(f"Carpeta purgada: {run_dir}")
        return retained_keys, notes

    retained_keys["run_artifacts_dir"] = str(run_dir)
    notes.append(
        "Modo final_only: el reporte final se genero correctamente, pero no fue posible purgar "
        "la carpeta de artefactos intermedios."
    )
    notes.append(
        "La ruta sigue disponible porque Windows/OneDrive probablemente mantuvo un bloqueo temporal "
        "sobre el directorio durante la fase de limpieza."
    )
    notes.append(f"Carpeta pendiente de purga: {run_dir}")
    return retained_keys, notes


def build_render_flavor_messages() -> list[tuple[str, str, str]]:
    return [
        ("Session", "Inicializando secuencia de ensamblaje del reporte final...", "info"),
        ("Manifest", "Persistiendo artefactos tecnicos y preparando vistas de salida...", "info"),
        ("Workbook", "Sincronizando hojas, tablas, enlaces y analitica ejecutiva...", "info"),
        ("Delivery", "Materializando workbook tecnico listo para revision y entrega...", "info"),
    ]


def save_report_outputs(
    settings: Settings,
    discovery_raw_nodes: list[dict],
    detail_raw_nodes: list[dict],
    risk_inventory: list[dict],
    parser_inventory: list[dict],
    parsed_risks: list[dict],
    attack_paths: list[dict],
    audit: AuditCollector,
) -> tuple[dict, list[str]]:
    """Persist all artifacts and generate the final technical workbook."""
    output_dir = ensure_output_dir(settings.output_dir)
    timestamp = get_timestamp()
    base_name = f"{settings.report_name}_{timestamp}"
    run_dir = ensure_output_dir(output_dir / "runs" / base_name)
    final_report_name = (
        f"Identity_Protection_Report_"
        f"{sanitize_filename_component(settings.target_domain)}_"
        f"{timestamp}_FINAL.xlsx"
    )
    final_report_path = output_dir.parent / final_report_name
    correlated_parsed_risks = enrich_parsed_risks_with_entity_correlation(parsed_risks)
    enriched_parsed_risks = enrich_parsed_risks_with_actionability(correlated_parsed_risks)

    unknown_rows = audit.build_unknown_risk_rows()
    parser_error_rows = audit.build_parser_error_rows()
    structure_issue_rows = audit.build_structure_issue_rows()
    error_rows = parser_error_rows + structure_issue_rows
    raw_sample_rows = audit.build_raw_samples_rows()
    raw_samples_json = audit.export_raw_samples_json()

    print_section(
        "Render Final",
        "Compilando artefactos y forjando el reporte tecnico final",
    )
    for label, message, tone in build_render_flavor_messages():
        print_status(label, message, tone)

    output_files = {
        "run_artifacts_dir": str(run_dir),
        "discovery_raw_json": str(run_dir / f"{base_name}_discovery_raw.json"),
        "detail_raw_json": str(run_dir / f"{base_name}_detail_raw.json"),
        "risk_inventory_csv": str(run_dir / f"{base_name}_risk_inventory.csv"),
        "parser_inventory_csv": str(run_dir / f"{base_name}_parser_inventory.csv"),
        "parsed_risks_csv": str(run_dir / f"{base_name}_parsed_risks.csv"),
        "attack_paths_csv": str(run_dir / f"{base_name}_attack_paths.csv"),
        "unknown_risk_types_csv": str(run_dir / f"{base_name}_unknown_risk_types.csv"),
        "parser_errors_csv": str(run_dir / f"{base_name}_parser_errors.csv"),
        "structure_issues_csv": str(run_dir / f"{base_name}_structure_issues.csv"),
        "raw_samples_csv": str(run_dir / f"{base_name}_raw_samples_overview.csv"),
        "raw_samples_json": str(run_dir / f"{base_name}_raw_samples.json"),
        "technical_excel": str(final_report_path),
    }

    save_json(discovery_raw_nodes, output_files["discovery_raw_json"])
    save_json(detail_raw_nodes, output_files["detail_raw_json"])
    save_csv(risk_inventory, output_files["risk_inventory_csv"], RISK_INVENTORY_FIELDS)
    save_csv(
        parser_inventory,
        output_files["parser_inventory_csv"],
        PARSER_INVENTORY_FIELDS,
    )
    save_csv(enriched_parsed_risks, output_files["parsed_risks_csv"], PARSED_RISK_FIELDS)
    save_csv(attack_paths, output_files["attack_paths_csv"], ATTACK_PATH_FIELDS)
    save_csv(
        unknown_rows,
        output_files["unknown_risk_types_csv"],
        UNKNOWN_RISK_FIELDS,
    )
    save_csv(parser_error_rows, output_files["parser_errors_csv"], ERROR_FIELDS)
    save_csv(
        structure_issue_rows,
        output_files["structure_issues_csv"],
        ERROR_FIELDS,
    )
    save_csv(raw_sample_rows, output_files["raw_samples_csv"], RAW_SAMPLE_FIELDS)
    save_json(raw_samples_json, output_files["raw_samples_json"])
    print_status(
        "Save Point",
        "Artefactos de corrida persistidos. Iniciando ensamblaje del workbook...",
        "success",
    )

    warnings = []
    excel_error = save_technical_excel(
        filepath=Path(output_files["technical_excel"]),
        settings=settings,
        risk_summary_rows=parser_inventory,
        parsed_risks=enriched_parsed_risks,
        attack_paths=attack_paths,
        unknown_rows=unknown_rows,
        error_rows=error_rows,
    )
    if excel_error:
        warnings.append(excel_error)
        output_files.pop("technical_excel", None)
        output_files["artifact_mode"] = settings.artifact_mode
        output_files["run_artifacts_dir"] = str(run_dir)
    else:
        print_status(
            "Quest Clear",
            "Workbook final generado. Aplicando politica de retencion de artefactos...",
            "success",
        )
        output_files, retention_notes = apply_artifact_retention(
            artifact_mode=settings.artifact_mode,
            run_dir=run_dir,
            output_files=output_files,
        )
        warnings.extend(retention_notes)

    return output_files, warnings


def save_technical_excel(
    filepath: Path,
    settings: Settings,
    risk_summary_rows: list[dict],
    parsed_risks: list[dict],
    attack_paths: list[dict],
    unknown_rows: list[dict],
    error_rows: list[dict],
) -> str | None:
    """Generate the final Excel workbook with executive and technical sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return (
            "No se genero el Excel tecnico porque falta la dependencia "
            "'openpyxl'."
        )

    workbook = Workbook()
    workbook.properties.creator = PROJECT_SIGNATURE
    workbook.properties.lastModifiedBy = PROJECT_SIGNATURE
    workbook.properties.title = "Identity Protection Technical Report"
    entity_correlation_rows = build_entity_correlation_rows(parsed_risks)
    lifecycle_correlation_rows = build_entity_correlation_rows(
        parsed_risks,
        family_filter=RISK_GROUP_SHEETS["Ciclo Vida Acceso"],
    )
    print_status(
        "Render Engine",
        "Sincronizando hojas principales, correlaciones y vistas tecnicas...",
        "info",
    )

    executive_sheet = workbook.active
    executive_sheet.title = SHEET_TITLES["executive"]
    chart_data_sheet = workbook.create_sheet(title="_chart_data")
    chart_data_sheet.sheet_state = "hidden"

    sheet_specs = [
        (SHEET_TITLES["risk_summary"], risk_summary_rows, PARSER_INVENTORY_FIELDS),
        (
            SHEET_TITLES["entity_correlation"],
            entity_correlation_rows,
            ENTITY_CORRELATION_FIELDS,
        ),
        (
            SHEET_TITLES["lifecycle_correlation"],
            lifecycle_correlation_rows,
            ENTITY_CORRELATION_FIELDS,
        ),
        (SHEET_TITLES["attack_paths"], attack_paths, ATTACK_PATH_FIELDS),
        (
            SHEET_TITLES["parsed_risks"],
            sort_parsed_risks(parsed_risks),
            PARSED_RISK_EXCEL_FIELDS,
        ),
    ]

    for title, rows in build_grouped_parsed_risk_sheets(parsed_risks).items():
        sheet_specs.append((title, rows, PARSED_RISK_EXCEL_FIELDS))

    sheet_specs.extend(
        [
            (SHEET_TITLES["unknown"], unknown_rows, UNKNOWN_RISK_FIELDS),
            (SHEET_TITLES["errors"], error_rows, ERROR_FIELDS),
        ]
    )

    sheets = {}
    for title, rows, fieldnames in sheet_specs:
        print_status("Sheet", f"Renderizando hoja: {title}", "info")
        worksheet = workbook.create_sheet(title=title)
        sheets[title] = worksheet
        write_sheet(
            worksheet,
            rows,
            fieldnames,
            Font,
            PatternFill,
            Alignment,
            Border,
            Side,
        )

    build_executive_summary_sheet(
        executive_sheet=executive_sheet,
        settings=settings,
        risk_summary_rows=risk_summary_rows,
        parsed_risks=parsed_risks,
        attack_paths=attack_paths,
        unknown_rows=unknown_rows,
        error_rows=error_rows,
        chart_data_sheet=chart_data_sheet,
        font_cls=Font,
        fill_cls=PatternFill,
        alignment_cls=Alignment,
        border_cls=Border,
        side_cls=Side,
        bar_chart_cls=BarChart,
        data_label_list_cls=DataLabelList,
        reference_cls=Reference,
    )

    color_sheet_tabs(workbook)

    workbook.save(filepath)
    print_status(
        "Archive",
        f"Workbook sellado en disco: {filepath.name}",
        "success",
    )
    return None


def write_sheet(
    worksheet,
    rows: list[dict],
    fieldnames: list[str],
    font_cls,
    fill_cls,
    alignment_cls,
    border_cls,
    side_cls,
) -> None:
    """Write one tabular worksheet with formatting, filters and hyperlinks."""
    effective_fieldnames = select_visible_fields(rows, fieldnames)
    header_fill = fill_cls(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    border = border_cls(
        left=side_cls(style="thin", color="D9E2F3"),
        right=side_cls(style="thin", color="D9E2F3"),
        top=side_cls(style="thin", color="D9E2F3"),
        bottom=side_cls(style="thin", color="D9E2F3"),
    )
    worksheet.append([FIELD_LABELS.get(field, field) for field in effective_fieldnames])

    for cell in worksheet[1]:
        cell.font = font_cls(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = alignment_cls(horizontal="center", vertical="center")
        cell.border = border

    for row in rows:
        worksheet.append(
            [excel_cell_value(field, row.get(field, "")) for field in effective_fieldnames]
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = alignment_cls(vertical="top", wrap_text=True)
            cell.border = border

        apply_attention_fills(row, effective_fieldnames, fill_cls, font_cls)

        status_index = next(
            (
                index
                for index, name in enumerate(effective_fieldnames, start=1)
                if name == "parse_status"
            ),
            None,
        )
        if status_index:
            apply_status_fill(row[status_index - 1], fill_cls)

    apply_column_formats(worksheet, effective_fieldnames, alignment_cls)

    for column_cells in worksheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        max_length = max(len(value) for value in values) if values else 0
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_length + 2, 12),
            60,
        )


def select_visible_fields(rows: list[dict], fieldnames: list[str]) -> list[str]:
    if not rows:
        return fieldnames

    visible = []
    for field in fieldnames:
        if any(row.get(field) not in ("", None) for row in rows):
            visible.append(field)

    return visible or fieldnames


def apply_status_fill(cell, fill_cls) -> None:
    status_colors = {
        "parsed": "E2F0D9",
        "parsed_empty": "D9EAF7",
        "requires_review": "FFF2CC",
        "structure_issue": "FCE4D6",
        "error": "F4CCCC",
    }
    fill_color = status_colors.get(str(cell.value or "").lower())
    if fill_color:
        cell.fill = fill_cls(fill_type="solid", start_color=fill_color, end_color=fill_color)


def apply_attention_fills(row, fieldnames: list[str], fill_cls, font_cls) -> None:
    attention_fields = {
        "severity": apply_severity_fill,
        "highest_severity": apply_severity_fill,
        "review_priority": apply_priority_fill,
        "actionability_level": apply_actionability_fill,
    }
    for index, fieldname in enumerate(fieldnames):
        painter = attention_fields.get(fieldname)
        if painter:
            painter(row[index], fill_cls, font_cls)


def apply_severity_fill(cell, fill_cls, font_cls) -> None:
    severity = str(cell.value or "").upper()
    palette = {
        "LOW": ("E2F0D9", "385723"),
        "MEDIUM": ("FFF2CC", "7F6000"),
        "HIGH": ("FCE4D6", "9C0006"),
        "CRITICAL": ("F4CCCC", "7F0000"),
    }
    style = palette.get(severity)
    if style:
        fill_color, font_color = style
        cell.fill = fill_cls(fill_type="solid", start_color=fill_color, end_color=fill_color)
        cell.font = font_cls(color=font_color, bold=True)


def apply_priority_fill(cell, fill_cls, font_cls) -> None:
    priority = str(cell.value or "").upper()
    palette = {
        "P1": ("F4CCCC", "7F0000"),
        "P2": ("FFF2CC", "7F6000"),
        "P3": ("DDEBF7", "1F4E78"),
    }
    style = palette.get(priority)
    if style:
        fill_color, font_color = style
        cell.fill = fill_cls(fill_type="solid", start_color=fill_color, end_color=fill_color)
        cell.font = font_cls(color=font_color, bold=True)


def apply_actionability_fill(cell, fill_cls, font_cls) -> None:
    actionability = str(cell.value or "")
    palette = {
        "Investigacion guiada": ("FCE4D6", "9C0006"),
        "Validacion en Falcon": ("FFF2CC", "7F6000"),
        "Accion directa": ("E2F0D9", "385723"),
    }
    style = palette.get(actionability)
    if style:
        fill_color, font_color = style
        cell.fill = fill_cls(fill_type="solid", start_color=fill_color, end_color=fill_color)
        cell.font = font_cls(color=font_color, bold=True)


def excel_cell_value(fieldname: str, value):
    if fieldname in PERCENTAGE_FIELDS and value not in ("", None):
        return float(value) / 100
    return value


def apply_column_formats(worksheet, fieldnames: list[str], alignment_cls) -> None:
    for column_index, fieldname in enumerate(fieldnames, start=1):
        for cell in worksheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=2,
            max_row=worksheet.max_row,
        ):
            for current in cell:
                if fieldname in PERCENTAGE_FIELDS and isinstance(current.value, (int, float)):
                    current.number_format = "0.00%"
                    current.alignment = alignment_cls(horizontal="right", vertical="top")
                elif fieldname in COUNT_FIELDS and isinstance(current.value, (int, float)):
                    current.number_format = "#,##0"
                    current.alignment = alignment_cls(horizontal="right", vertical="top")
                elif fieldname in URL_FIELDS and current.value:
                    current.hyperlink = str(current.value)
                    current.style = "Hyperlink"


def build_executive_summary_sheet(
    executive_sheet,
    settings: Settings,
    risk_summary_rows: list[dict],
    parsed_risks: list[dict],
    attack_paths: list[dict],
    unknown_rows: list[dict],
    error_rows: list[dict],
    chart_data_sheet,
    font_cls,
    fill_cls,
    alignment_cls,
    border_cls,
    side_cls,
    bar_chart_cls,
    data_label_list_cls,
    reference_cls,
) -> None:
    """Render the executive cover sheet using precomputed analytics."""
    analytics = build_executive_analytics(
        risk_summary_rows=risk_summary_rows,
        parsed_risks=parsed_risks,
        attack_paths=attack_paths,
        unknown_rows=unknown_rows,
        error_rows=error_rows,
    )
    metrics = analytics["metrics"]

    executive_sheet.sheet_view.showGridLines = False
    executive_sheet.freeze_panes = "A9"

    title_fill = fill_cls(fill_type="solid", start_color="163A5F", end_color="163A5F")
    accent_fill = fill_cls(fill_type="solid", start_color="DCE6F1", end_color="DCE6F1")
    card_fill = fill_cls(fill_type="solid", start_color="EAF2F8", end_color="EAF2F8")
    section_fill = fill_cls(fill_type="solid", start_color="5B9BD5", end_color="5B9BD5")
    soft_fill = fill_cls(fill_type="solid", start_color="F6F9FC", end_color="F6F9FC")
    border = border_cls(
        left=side_cls(style="thin", color="B8CCE4"),
        right=side_cls(style="thin", color="B8CCE4"),
        top=side_cls(style="thin", color="B8CCE4"),
        bottom=side_cls(style="thin", color="B8CCE4"),
    )

    executive_sheet.merge_cells("A1:H2")
    title_cell = executive_sheet["A1"]
    title_cell.value = "Reporte Tecnico de Identity Protection"
    title_cell.font = font_cls(size=16, bold=True, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = alignment_cls(horizontal="center", vertical="center")

    executive_sheet.merge_cells("A3:H3")
    subtitle_cell = executive_sheet["A3"]
    subtitle_cell.value = (
        f"Dominio analizado: {settings.target_domain} | "
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    subtitle_cell.font = font_cls(size=10, italic=True, color="1F1F1F")
    subtitle_cell.fill = accent_fill
    subtitle_cell.alignment = alignment_cls(horizontal="center", vertical="center")

    kpis = [
        ("Riesgos detectados", f"{metrics['total_risks']:,}"),
        ("Tipos de riesgo", f"{metrics['total_risk_types']:,}"),
        ("Familia dominante", metrics["lead_family"]),
        ("Riesgo principal", metrics["lead_risk_title"]),
        ("Top 3 concentran", f"{metrics['top3_share']:.2f}%"),
        ("Rutas de ataque", f"{metrics['attack_path_count']:,}"),
    ]

    card_ranges = ["A5:B7", "C5:D7", "E5:F7", "G5:H7", "I5:J7", "K5:L7"]
    for (label, value), cell_range in zip(kpis, card_ranges):
        executive_sheet.merge_cells(cell_range)
        cell = executive_sheet[cell_range.split(":")[0]]
        cell.value = f"{label}\n{value}"
        cell.font = font_cls(size=12, bold=True, color="163A5F")
        cell.fill = card_fill
        cell.alignment = alignment_cls(horizontal="center", vertical="center", wrap_text=True)
        for row in executive_sheet[cell_range]:
            for current in row:
                current.border = border

    make_section_header(
        executive_sheet,
        "A10:F10",
        "Lectura Ejecutiva",
        font_cls,
        section_fill,
        alignment_cls,
        border,
    )
    make_section_header(
        executive_sheet,
        "G10:L10",
        "Acciones Sugeridas",
        font_cls,
        section_fill,
        alignment_cls,
        border,
    )
    write_bullet_box(
        executive_sheet,
        "A11:F16",
        analytics["key_findings"] + analytics["operational_highlights"][:1],
        font_cls,
        soft_fill,
        alignment_cls,
        border,
    )
    write_bullet_box(
        executive_sheet,
        "G11:L16",
        analytics["suggested_actions"],
        font_cls,
        soft_fill,
        alignment_cls,
        border,
    )

    make_section_header(
        executive_sheet,
        "A18:F18",
        "Top 5 Riesgos",
        font_cls,
        section_fill,
        alignment_cls,
        border,
    )
    make_section_header(
        executive_sheet,
        "G18:L18",
        "Familias de Riesgo",
        font_cls,
        section_fill,
        alignment_cls,
        border,
    )

    write_helper_table(
        executive_sheet,
        start_row=19,
        start_col=1,
        headers=["Riesgo", "Cantidad", "Porcentaje"],
        rows=[
            [
                row.get("risk_title", row.get("risk_type", "")),
                row.get("count", 0),
                float(row.get("percentage", 0)) / 100,
            ]
            for row in analytics["top_risks"]
        ],
        fill_cls=fill_cls,
        font_cls=font_cls,
        alignment_cls=alignment_cls,
        border=border,
    )
    write_helper_table(
        executive_sheet,
        start_row=19,
        start_col=7,
        headers=["Familia", "Cantidad"],
        rows=[[family, count] for family, count in analytics["condensed_families"]] or [["Sin datos", 0]],
        fill_cls=fill_cls,
        font_cls=font_cls,
        alignment_cls=alignment_cls,
        border=border,
    )

    write_chart_data_table(
        chart_data_sheet,
        start_row=19,
        start_col=1,
        headers=["Etiqueta", "Cantidad"],
        rows=list(zip(analytics["top_risk_chart_labels"], [row["count"] for row in analytics["top_risks"]])),
    )
    write_chart_data_table(
        chart_data_sheet,
        start_row=19,
        start_col=4,
        headers=["Etiqueta", "Cantidad"],
        rows=list(zip(analytics["family_chart_labels"], [count for _, count in analytics["condensed_families"]])),
    )

    add_top_risks_chart(
        executive_sheet,
        chart_data_sheet=chart_data_sheet,
        anchor="A27",
        title="Top 5 riesgos por cantidad",
        reference_cls=reference_cls,
        bar_chart_cls=bar_chart_cls,
        data_label_list_cls=data_label_list_cls,
        rows=len(analytics["top_risks"]),
        max_label_length=max((len(label) for label in analytics["top_risk_chart_labels"]), default=10),
    )
    add_family_distribution_chart(
        executive_sheet,
        chart_data_sheet=chart_data_sheet,
        anchor="H27",
        title="Top familias por exposicion",
        reference_cls=reference_cls,
        bar_chart_cls=bar_chart_cls,
        data_label_list_cls=data_label_list_cls,
        rows=len(analytics["condensed_families"]),
        max_label_length=max((len(label) for label in analytics["family_chart_labels"]), default=10),
    )

    make_section_header(
        executive_sheet,
        "A49:L49",
        "Nota Metodologica",
        font_cls,
        section_fill,
        alignment_cls,
        border,
    )
    executive_sheet.merge_cells("A50:L53")
    note_cell = executive_sheet["A50"]
    note_cell.value = (
        "Esta portada ejecutiva resume hallazgos a partir de los conteos, familias, "
        "rutas de ataque y resultados de parseo ya presentes en el workbook. "
        "No agrega fuentes externas ni replica evidencia sensible mas alla de los "
        "agregados necesarios para presentacion."
    )
    note_cell.alignment = alignment_cls(wrap_text=True, vertical="top")
    note_cell.fill = accent_fill
    note_cell.border = border

    executive_sheet.column_dimensions["A"].width = 28
    executive_sheet.column_dimensions["B"].width = 12
    executive_sheet.column_dimensions["C"].width = 12
    executive_sheet.column_dimensions["D"].width = 12
    executive_sheet.column_dimensions["E"].width = 12
    executive_sheet.column_dimensions["F"].width = 22
    executive_sheet.column_dimensions["G"].width = 12
    executive_sheet.column_dimensions["H"].width = 18
    executive_sheet.column_dimensions["I"].width = 12
    executive_sheet.column_dimensions["J"].width = 18
    executive_sheet.column_dimensions["K"].width = 12
    executive_sheet.column_dimensions["L"].width = 12


def write_helper_table(
    worksheet,
    start_row: int,
    start_col: int,
    headers: list[str],
    rows: list[list],
    fill_cls,
    font_cls,
    alignment_cls,
    border,
) -> None:
    header_fill = fill_cls(fill_type="solid", start_color="5B9BD5", end_color="5B9BD5")

    for offset, header in enumerate(headers):
        cell = worksheet.cell(row=start_row, column=start_col + offset, value=header)
        cell.font = font_cls(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = alignment_cls(horizontal="center")
        cell.border = border

    for row_offset, data_row in enumerate(rows, start=1):
        for col_offset, value in enumerate(data_row):
            cell = worksheet.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)
            cell.alignment = alignment_cls(vertical="top", wrap_text=True)
            cell.border = border
            if isinstance(value, float) and headers[col_offset] == "Porcentaje":
                cell.number_format = "0.00%"
            elif isinstance(value, int):
                cell.number_format = "#,##0"


def add_top_risks_chart(
    worksheet,
    chart_data_sheet,
    anchor: str,
    title: str,
    reference_cls,
    bar_chart_cls,
    data_label_list_cls,
    rows: int,
    max_label_length: int,
) -> None:
    chart = bar_chart_cls()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "Riesgo"
    chart.x_axis.title = "Cantidad"
    chart.height, chart.width = get_chart_dimensions(rows, max_label_length)
    data = reference_cls(chart_data_sheet, min_col=2, min_row=19, max_row=19 + rows)
    categories = reference_cls(chart_data_sheet, min_col=1, min_row=20, max_row=19 + rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend = None
    chart.gapWidth = 35
    chart.dLbls = data_label_list_cls()
    chart.dLbls.showVal = False
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    chart.dLbls.showLegendKey = False
    chart.x_axis.title = None
    chart.y_axis.title = None
    worksheet.add_chart(chart, anchor)


def add_family_distribution_chart(
    worksheet,
    chart_data_sheet,
    anchor: str,
    title: str,
    reference_cls,
    bar_chart_cls,
    data_label_list_cls,
    rows: int,
    max_label_length: int,
) -> None:
    chart = bar_chart_cls()
    chart.type = "bar"
    chart.style = 11
    chart.title = title
    chart.height, chart.width = get_chart_dimensions(rows, max_label_length)
    data = reference_cls(chart_data_sheet, min_col=5, min_row=19, max_row=19 + rows)
    labels = reference_cls(chart_data_sheet, min_col=4, min_row=20, max_row=19 + rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.legend = None
    chart.gapWidth = 35
    chart.dLbls = data_label_list_cls()
    chart.dLbls.showVal = False
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    chart.dLbls.showLegendKey = False
    chart.x_axis.title = None
    chart.y_axis.title = None
    worksheet.add_chart(chart, anchor)


def count_family_rows(worksheet, start_col: int = 6, start_row: int = 11) -> int:
    count = 0
    row = start_row
    while worksheet.cell(row=row, column=start_col).value:
        count += 1
        row += 1
    return count


def get_chart_dimensions(rows: int, max_label_length: int) -> tuple[float, float]:
    chart_height = max(8.5, min(13.0, 6.5 + (rows * 1.0)))
    chart_width = max(11.5, min(16.0, 9.5 + (max_label_length * 0.16)))
    return chart_height, chart_width


def write_chart_data_table(
    worksheet,
    start_row: int,
    start_col: int,
    headers: list[str],
    rows: list[tuple[str, int]],
) -> None:
    for offset, header in enumerate(headers):
        worksheet.cell(row=start_row, column=start_col + offset, value=header)

    for row_offset, (label, value) in enumerate(rows, start=1):
        worksheet.cell(row=start_row + row_offset, column=start_col, value=label)
        worksheet.cell(row=start_row + row_offset, column=start_col + 1, value=value)


def make_section_header(
    worksheet,
    cell_range: str,
    title: str,
    font_cls,
    fill,
    alignment_cls,
    border,
) -> None:
    worksheet.merge_cells(cell_range)
    start_cell = worksheet[cell_range.split(":")[0]]
    start_cell.value = title
    start_cell.font = font_cls(bold=True, color="FFFFFF")
    start_cell.fill = fill
    start_cell.alignment = alignment_cls(horizontal="left", vertical="center")
    for row in worksheet[cell_range]:
        for cell in row:
            cell.border = border


def write_bullet_box(
    worksheet,
    cell_range: str,
    lines: list[str],
    font_cls,
    fill,
    alignment_cls,
    border,
) -> None:
    worksheet.merge_cells(cell_range)
    start_cell = worksheet[cell_range.split(":")[0]]
    start_cell.value = "\n".join(f"- {line}" for line in lines if line)
    start_cell.font = font_cls(size=10, color="1F1F1F")
    start_cell.fill = fill
    start_cell.alignment = alignment_cls(vertical="top", wrap_text=True)
    for row in worksheet[cell_range]:
        for cell in row:
            cell.border = border


def color_sheet_tabs(workbook) -> None:
    tab_colors = {
        SHEET_TITLES["executive"]: "1F4E78",
        SHEET_TITLES["risk_summary"]: "2F75B5",
        SHEET_TITLES["parsed_risks"]: "70AD47",
        SHEET_TITLES["attack_paths"]: "C55A11",
        SHEET_TITLES["unknown"]: "BF9000",
        SHEET_TITLES["errors"]: "C00000",
    }
    for worksheet in workbook.worksheets:
        worksheet.sheet_properties.tabColor = tab_colors.get(worksheet.title, "5B9BD5")
